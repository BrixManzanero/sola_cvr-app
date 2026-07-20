"""
backend/tiktok.py
TikTok Shop Analytics CVR extractor.

CVR formula: Orders ÷ Visitors   (unified across all platforms)
  – Numerator   = Orders (order count), summed from the daily rows
  – Denominator = Visitors (unique store visitors)
  – TikTok's own native CVR (Customers ÷ Visitors) is still aggregated and
    shown in the audit panel for reference only — it is NOT the reported CVR.

Always exports daily rows — aggregated to monthly here.
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.utils import (
    fmt_cvr, fmt_num,
    safe_read, safe_xl,
    make_excel_styles, to_xlsx_bytes,
    xl_section, xl_header, xl_row,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

REQUIRED_COLS   = {"Conversion rate", "Page views", "Visitors", "Orders", "Product clicks"}
DATE_MARKER     = "Analysis date:"
DAILY_MARKER    = "Daily data"
DISPLAY_HEADERS = [
    "Month / Period", "Page Views", "Visitors",
    "Product Clicks", "Orders", "Conversion Rate",
]


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────

def validate(uploaded_file) -> tuple:
    if not uploaded_file.name.lower().endswith(".xlsx"):
        return False, "Expected a .xlsx file"
    try:
        xl = safe_xl(uploaded_file)
    except Exception as exc:
        return False, f"Cannot open file: {exc}"
    if "Sheet1" not in xl.sheet_names:
        return False, f"No 'Sheet1' found — sheets: {xl.sheet_names}"
    try:
        df   = safe_read(uploaded_file, sheet_name="Sheet1", header=None)
        a1   = str(df.iloc[0, 0])
        if DATE_MARKER not in a1:
            return False, f"Cell A1 missing '{DATE_MARKER}' — not a TikTok export"
        flat = df.iloc[:, 0].tolist()
        if not any(DAILY_MARKER in str(v) for v in flat):
            return False, f"No '{DAILY_MARKER}' section found"
        av   = {str(v) for v in df.values.flatten() if pd.notna(v) and str(v).strip()}
        miss = REQUIRED_COLS - av
        if miss:
            return False, f"Missing columns: {miss}"
    except Exception as exc:
        return False, f"Error reading sheet: {exc}"
    period = a1.replace(DATE_MARKER, "").split("Comparison")[0].strip()
    return True, period


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────

def extract(uploaded_file) -> dict:
    """
    Read one TikTok file and return monthly CVR data.
    CVR = Customers ÷ Visitors per month.
    Also stores raw Customers + Visitors in monthly DataFrame
    so the auditor can cross-check against TikTok's summary CVR.
    """
    df = safe_read(uploaded_file, sheet_name="Sheet1", header=None)

    # Locate overview and daily sections
    overview_hdr = None
    daily_start  = None
    for i, val in enumerate(df.iloc[:, 0].tolist()):
        if "Data overview" in str(val): overview_hdr = i + 1
        if DAILY_MARKER    in str(val): daily_start  = i + 1

    # Summary row
    ov_cols = list(df.iloc[overview_hdr])
    sum_raw = dict(zip(ov_cols, df.iloc[overview_hdr + 1].tolist()))

    a1     = str(df.iloc[0, 0])
    period = a1.replace(DATE_MARKER, "").split("Comparison")[0].strip()
    try:
        start_dt = datetime.strptime(period.split("-")[0].strip(), "%d/%m/%Y")
    except Exception:
        start_dt = datetime.min

    # TikTok's OWN summary CVR (Customers ÷ Visitors) — kept only for reference/audit.
    source_cvr_raw = sum_raw.get("Conversion rate", 0)

    # NEW unified CVR = Orders ÷ Visitors (comparable across all platforms).
    def _to_f(v):
        try:
            return float(str(v).replace(",", "").replace("%", "").strip())
        except Exception:
            return 0.0
    _sum_ord = _to_f(sum_raw.get("Orders", 0))
    _sum_vis = _to_f(sum_raw.get("Visitors", 0))
    summary_cvr = f"{_sum_ord / _sum_vis * 100:.2f}%" if _sum_vis else "N/A"

    summary = {
        "Period":          period,
        "Page Views":      fmt_num(sum_raw.get("Page views", 0)),
        "Visitors":        fmt_num(sum_raw.get("Visitors", 0)),
        "Product Clicks":  fmt_num(sum_raw.get("Product clicks", 0)),
        "Orders":          fmt_num(sum_raw.get("Orders", 0)),
        "Conversion Rate": summary_cvr,   # Orders ÷ Visitors
    }

    # Daily rows
    daily_cols = list(df.iloc[daily_start])
    daily_df   = df.iloc[daily_start + 1:].copy()
    daily_df.columns = daily_cols
    daily_df   = daily_df.dropna(subset=["Date"]).reset_index(drop=True)

    # NEW unified CVR = Orders ÷ Visitors.
    # 'Customers' (unique buyers) is still aggregated when present, but only for
    # reference/audit — it is NOT used as the CVR numerator anymore.
    use_customers = "Customers" in daily_df.columns
    cvr_warnings  = []

    num_cols = ["Page views", "Visitors", "Product clicks", "Orders"]
    if use_customers:
        num_cols.append("Customers")

    for col in num_cols:
        daily_df[col] = pd.to_numeric(daily_df[col], errors="coerce").fillna(0)

    daily_df["_date"]  = pd.to_datetime(daily_df["Date"], dayfirst=True, errors="coerce")
    daily_df           = daily_df.dropna(subset=["_date"])
    daily_df["_month"] = daily_df["_date"].dt.to_period("M")

    agg_dict = {
        "PV":     ("Page views",     "sum"),
        "Vis":    ("Visitors",       "sum"),
        "Clk":    ("Product clicks", "sum"),
        "Ord":    ("Orders",         "sum"),
        "DayCnt": ("_date",          "count"),   # actual daily rows per month
        "MinDt":  ("_date",          "min"),     # first date in this month's data
    }
    if use_customers:
        agg_dict["Cust"] = ("Customers", "sum")

    grouped = daily_df.groupby("_month").agg(**agg_dict).reset_index()

    # CVR: Orders ÷ Visitors  (unified cross-platform formula)
    num_series = grouped["Ord"]
    grouped["CVR"] = (
        num_series / grouped["Vis"].replace(0, float("nan")) * 100
    ).round(2)
    grouped["CVR"] = grouped["CVR"].apply(
        lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A"
    )

    rows = []
    for _, row in grouped.iterrows():
        dt = row["_month"].to_timestamp()
        r  = {
            "Date":            dt,
            "Month Label":     dt.strftime("%B %Y"),
            "Page Views":      fmt_num(row["PV"]),
            "Visitors":        fmt_num(row["Vis"]),
            "Product Clicks":  fmt_num(row["Clk"]),
            "Orders":          fmt_num(row["Ord"]),
            "Conversion Rate": row["CVR"],
            "_day_count":      int(row["DayCnt"]),
            "_period_start":   row["MinDt"].to_pydatetime() if hasattr(row["MinDt"], "to_pydatetime") else row["MinDt"],
            "_period_end":     dt.replace(day=1) + pd.offsets.MonthEnd(0),  # last day of month
        }
        # Keep raw Customers + Visitors for auditor verification
        if use_customers:
            r["Customers"] = fmt_num(row["Cust"])
        rows.append(r)

    return {
        "summary":          summary,
        "monthly":          pd.DataFrame(rows),
        "start_date":       start_dt,
        "name":             uploaded_file.name,
        "source_cvr_raw":   source_cvr_raw,   # passed to auditor
        "cvr_warnings":     cvr_warnings,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(monthly_df: pd.DataFrame, summaries: list) -> bytes:
    """
    Generate a plain TikTok CVR Excel — single 'Traffic Conversion' tab:
        Month | Page Views | Visitors | Product Clicks | Orders | CVR  + TOTAL row
    No colors or fills. TOTAL CVR = total Orders ÷ total Visitors (unified formula).
    """
    from backend.utils import write_plain_sheet

    def _num(v) -> float:
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    wb = Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("Traffic Conversion")
    ws.sheet_view.showGridLines = True

    rows = []
    t_pv = t_vis = t_clk = t_ord = t_cust = 0.0
    has_cust = "Customers" in monthly_df.columns
    for _, row in monthly_df.iterrows():
        pv  = _num(row.get("Page Views", 0))
        vis = _num(row.get("Visitors", 0))
        clk = _num(row.get("Product Clicks", 0))
        ordr= _num(row.get("Orders", 0))
        t_pv += pv; t_vis += vis; t_clk += clk; t_ord += ordr
        row_vals = [
            row.get("Month Label", str(row["Date"])),
            f"{int(pv):,}", f"{int(vis):,}", f"{int(clk):,}", f"{int(ordr):,}",
        ]
        if has_cust:
            cust = _num(row.get("Customers", 0))
            t_cust += cust
            row_vals.append(f"{int(cust):,}")
        row_vals.append(row.get("Conversion Rate", "N/A"))
        rows.append(row_vals)

    # No aggregate TOTAL row — visitor de-dup basis differs across platforms.
    if has_cust:
        headers = ["Month", "Page Views", "Visitors", "Product Clicks", "Orders", "Customers", "CVR"]
        widths  = [14, 14, 12, 16, 10, 12, 12]
    else:
        headers = ["Month", "Page Views", "Visitors", "Product Clicks", "Orders", "CVR"]
        widths  = [14, 14, 12, 16, 10, 12]

    write_plain_sheet(ws, headers=headers, rows=rows, total_row=None, col_widths=widths)
    return to_xlsx_bytes(wb)
