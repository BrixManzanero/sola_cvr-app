"""
backend/lazada.py
Lazada Business Advisor CVR extractor.
Uses the summary row per file (official unique-visitor metrics).
Supports both .xls (xlrd) and .xlsx (openpyxl) formats.
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from backend.utils import (
    fmt_cvr, fmt_num,
    safe_read, safe_xl,
    make_excel_styles, to_xlsx_bytes,
    xl_section, xl_header, xl_row,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

LAZADA_MARKER   = "Data Source: Lazada - Business Advisor - Dashboard"
KEY_SHEET       = "Key Metrics"
REQUIRED_COLS   = {"Conversion Rate", "Visitors", "Buyers", "Orders", "Pageviews"}
DISPLAY_HEADERS = [
    "Month / Period", "Pageviews", "Visitors",
    "Buyers", "Orders", "Conversion Rate",
]


# ─────────────────────────────────────────────────────────────────────────────
# DATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def period_to_label(period: str) -> str:
    """
    Convert '2025-10-01~2025-10-31' to 'October 2025'.
    Multi-month ranges return 'Oct 2025 – Dec 2025'.
    """
    try:
        start_s, end_s = str(period).split("~")
        start = datetime.strptime(start_s.strip(), "%Y-%m-%d")
        end   = datetime.strptime(end_s.strip(),   "%Y-%m-%d")
        if start.month == end.month and start.year == end.year:
            return start.strftime("%B %Y")
        return f"{start.strftime('%b %Y')} – {end.strftime('%b %Y')}"
    except Exception:
        return str(period)


def period_start(period: str) -> datetime:
    """Return the start date of a Lazada period string for sorting."""
    try:
        return datetime.strptime(str(period).split("~")[0].strip(), "%Y-%m-%d")
    except Exception:
        return datetime.min


def _engine(filename: str) -> str:
    return "xlrd" if filename.lower().endswith(".xls") else "openpyxl"


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────

def validate(uploaded_file) -> tuple[bool, str]:
    """
    Returns (True, period_string) when the file is a valid Lazada export.
    Returns (False, reason_string) when it is not.
    """
    name = uploaded_file.name
    if not name.lower().endswith((".xls", ".xlsx")):
        return False, "Expected a .xls or .xlsx file"

    engine = _engine(name)
    try:
        xl = safe_xl(uploaded_file, engine=engine)
    except Exception as exc:
        return False, f"Cannot open file: {exc}"

    if KEY_SHEET not in xl.sheet_names:
        return False, f"No '{KEY_SHEET}' sheet found — sheets: {xl.sheet_names}"

    try:
        df   = safe_read(uploaded_file, sheet_name=KEY_SHEET, header=None, engine=engine)
        a1   = str(df.iloc[0, 0])

        if LAZADA_MARKER not in a1:
            return False, "Cell A1 missing Lazada marker — not a Lazada Business Advisor file"

        all_vals = {str(v) for v in df.values.flatten() if pd.notna(v) and str(v).strip()}
        missing  = REQUIRED_COLS - all_vals
        if missing:
            return False, f"Missing columns: {missing}"

    except Exception as exc:
        return False, f"Error reading sheet: {exc}"

    headers  = list(df.iloc[5])
    sum_row  = dict(zip(headers, df.iloc[6].tolist()))
    period   = str(sum_row.get("Date", ""))
    return True, period


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────

def extract(uploaded_file) -> dict:
    """
    Read the summary row from a Lazada file and return one monthly data point.
    The summary row provides Lazada's official deduplicated Visitors/Buyers/CVR.
    """
    engine  = _engine(uploaded_file.name)
    df      = safe_read(uploaded_file, sheet_name=KEY_SHEET, header=None, engine=engine)
    headers = list(df.iloc[5])
    sum_row = dict(zip(headers, df.iloc[6].tolist()))
    period  = str(sum_row.get("Date", ""))

    def _period_days(p: str) -> int:
        """Count days in '2026-05-01~2026-05-27' → 27."""
        try:
            s, e = str(p).split("~")
            sd = datetime.strptime(s.strip(), "%Y-%m-%d")
            ed = datetime.strptime(e.strip(), "%Y-%m-%d")
            return (ed - sd).days + 1
        except Exception:
            return 0

    def _period_end_lz(p: str) -> datetime:
        try:
            return datetime.strptime(str(p).split("~")[1].strip(), "%Y-%m-%d")
        except Exception:
            return datetime.min

    p_start = period_start(period)
    p_end   = _period_end_lz(period)
    day_cnt = _period_days(period)

    row = {
        "Date":            p_start,
        "Month Label":     period_to_label(period),
        "Pageviews":       fmt_num(sum_row.get("Pageviews", 0)),
        "Visitors":        fmt_num(sum_row.get("Visitors",  0)),
        "Buyers":          fmt_num(sum_row.get("Buyers",    0)),
        "Orders":          fmt_num(sum_row.get("Orders",    0)),
        "Conversion Rate": fmt_cvr(sum_row.get("Conversion Rate", 0)),
        "_day_count":      day_cnt,
        "_period_start":   p_start,
        "_period_end":     p_end,
    }

    return {
        "row":        row,
        "summary":    row.copy(),
        "period":     period,
        "start_date": p_start,
        "name":       uploaded_file.name,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(monthly_df: pd.DataFrame, summaries: list) -> bytes:
    """
    Generate a plain Lazada CVR Excel — single 'Traffic Conversion' tab:
        Month | Pageviews | Visitors | Buyers | Orders | CVR  + TOTAL row
    No colors or fills. TOTAL CVR = total buyers ÷ total visitors (Lazada definition).
    """
    from backend.utils import write_plain_sheet

    def _num(v) -> float:
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Traffic Conversion")
    ws.sheet_view.showGridLines = True

    rows = []
    t_pv = t_vis = t_buy = t_ord = 0.0
    for _, row in monthly_df.iterrows():
        pv  = _num(row.get("Pageviews", 0))
        vis = _num(row.get("Visitors", 0))
        buy = _num(row.get("Buyers", 0))
        ordr= _num(row.get("Orders", 0))
        t_pv += pv; t_vis += vis; t_buy += buy; t_ord += ordr
        rows.append([
            row["Month Label"],
            f"{int(pv):,}", f"{int(vis):,}", f"{int(buy):,}",
            f"{int(ordr):,}", row.get("Conversion Rate", "N/A"),
        ])

    total_cvr = f"{t_buy / t_vis * 100:.2f}%" if t_vis else "N/A"
    total_row = ["TOTAL", f"{int(t_pv):,}", f"{int(t_vis):,}",
                 f"{int(t_buy):,}", f"{int(t_ord):,}", total_cvr]

    write_plain_sheet(
        ws,
        headers=["Month", "Pageviews", "Visitors", "Buyers", "Orders", "CVR"],
        rows=rows,
        total_row=total_row,
        col_widths=[14, 14, 12, 10, 10, 12],
    )
    return to_xlsx_bytes(wb)
