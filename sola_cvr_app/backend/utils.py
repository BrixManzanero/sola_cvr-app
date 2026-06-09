"""
backend/utils.py
Shared helpers: number/CVR formatting, safe pandas readers, Excel style builders.
No Streamlit imports — pure Python only.
"""

import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# VALUE FORMATTERS
# ─────────────────────────────────────────────────────────────────────────────

def fmt_cvr(val) -> str:
    """
    Normalise a CVR value to a clean '2.66%' string.
    Handles: '2.66%' string | 0.0266 decimal float | 2.66 percentage float.
    """
    if val is None or str(val).strip() in ("", "-", "nan"):
        return "N/A"
    s = str(val).strip()
    if s.endswith("%"):
        return s
    try:
        f = float(s)
        if f < 1:          # Shopee/TikTok store CVR as decimal (0.0266)
            f *= 100
        return f"{f:.2f}%"
    except Exception:
        return "N/A"


def fmt_num(val) -> str:
    """Format a numeric value with comma separators e.g. 1212066 → '1,212,066'."""
    try:
        s = str(val).strip().replace(",", "")
        if s in ("-", "", "nan"):
            return "0"
        return f"{int(float(s)):,}"
    except Exception:
        return str(val)


# ─────────────────────────────────────────────────────────────────────────────
# SAFE PANDAS READERS  (reset file pointer before each read)
# ─────────────────────────────────────────────────────────────────────────────

def safe_read(f, **kwargs):
    """
    Read a pandas Excel from an uploaded file.
    Always creates a fresh BytesIO copy so that multiple reads of
    different sheets from the same UploadedFile object work correctly.
    """
    import io
    import pandas as pd
    f.seek(0)
    data = f.read()
    f.seek(0)
    return pd.read_excel(io.BytesIO(data), **kwargs)


def safe_xl(f, **kwargs):
    """
    Return a pandas ExcelFile for sheet-name inspection.
    Uses a fresh BytesIO copy — safe for Streamlit UploadedFile objects.
    """
    import io
    import pandas as pd
    f.seek(0)
    data = f.read()
    f.seek(0)
    return pd.ExcelFile(io.BytesIO(data), **kwargs)


# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def to_xlsx_bytes(wb) -> bytes:
    """Save an openpyxl Workbook to bytes for Streamlit download_button."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def write_plain_sheet(ws, headers: list, rows: list, total_row: list = None,
                      col_widths: list = None):
    """
    Write a plain, unstyled table to a worksheet.
    No colors, no fills, no borders. Headers are bold only; everything else plain.
      headers     : list of column titles
      rows        : list of row value-lists
      total_row   : optional list of values for a bold TOTAL row
      col_widths  : optional list of column widths
    """
    from openpyxl.styles import Font

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    bold = Font(bold=True)

    # Header row (bold text only, no fill/color)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = bold

    # Data rows
    r = 2
    for row_vals in rows:
        for c, v in enumerate(row_vals, 1):
            ws.cell(row=r, column=c, value=v)
        r += 1

    # Optional TOTAL row (bold)
    if total_row:
        for c, v in enumerate(total_row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = bold

    ws.freeze_panes = ws.cell(row=2, column=1)


def check_dupes(pairs: list) -> dict:
    """
    Return a dict of {month_label: count} for any month that
    appears more than once across all uploaded files.
    """
    from collections import Counter
    counts = Counter(m for m, _ in pairs)
    return {m: cnt for m, cnt in counts.items() if cnt > 1}


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL STYLE FACTORY
# ─────────────────────────────────────────────────────────────────────────────

def make_excel_styles(hdr_color: str, sec_color: str,
                      sec_text: str, title_color: str) -> dict:
    """
    Build a style dictionary used by all three platform Excel builders.
    Pass hex strings without '#' e.g. '1E5799'.
    """
    thin = Side(style="thin", color="B0C4DE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "hdr":   PatternFill("solid", start_color=hdr_color,   end_color=hdr_color),
        "sec":   PatternFill("solid", start_color=sec_color,   end_color=sec_color),
        "sum":   PatternFill("solid", start_color="F4A460",    end_color="F4A460"),
        "alt":   PatternFill("solid", start_color="EFF6FF",    end_color="EFF6FF"),
        "wht":   PatternFill("solid", start_color="FFFFFF",    end_color="FFFFFF"),
        "ftitl": Font(name="Arial", bold=True, color=title_color, size=13),
        "fhdr":  Font(name="Arial", bold=True, color="FFFFFF",    size=10),
        "fsec":  Font(name="Arial", bold=True, color=sec_text,    size=10),
        "fsum":  Font(name="Arial", bold=True, color="7B3F00",    size=10),
        "fbod":  Font(name="Arial", size=10),
        "brd":   border,
        "ctr":   Alignment(horizontal="center", vertical="center"),
        "lft":   Alignment(horizontal="left",   vertical="center"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL CELL WRITERS  (used by all platform excel builders)
# ─────────────────────────────────────────────────────────────────────────────

def xl_section(ws, row: int, text: str, s: dict, ncols: int):
    """Write a merged section-label row."""
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    ws[f"A{row}"] = text
    ws[f"A{row}"].font      = s["fsec"]
    ws[f"A{row}"].fill      = s["sec"]
    ws[f"A{row}"].alignment = s["lft"]
    ws[f"A{row}"].border    = s["brd"]
    ws.row_dimensions[row].height = 20


def xl_header(ws, row: int, headers: list, s: dict, extra_fills: list = None):
    """Write a header row; optionally override individual cell fills."""
    for c, h in enumerate(headers, 1):
        cell           = ws.cell(row=row, column=c, value=h)
        cell.font      = s["fhdr"]
        cell.fill      = extra_fills[c - 1] if extra_fills else s["hdr"]
        cell.alignment = s["lft"] if c == 1 else s["ctr"]
        cell.border    = s["brd"]
    ws.row_dimensions[row].height = 22


def xl_row(ws, row: int, vals: list, font, fill, s: dict):
    """Write a data row with uniform font and fill."""
    for c, val in enumerate(vals, 1):
        cell           = ws.cell(row=row, column=c, value=val)
        cell.font      = font
        cell.fill      = fill
        cell.alignment = s["lft"] if c == 1 else s["ctr"]
        cell.border    = s["brd"]
    ws.row_dimensions[row].height = 20
