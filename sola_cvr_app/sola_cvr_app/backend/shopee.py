"""
backend/shopee.py
Shopee Shop Stats CVR extractor.
Handles both monthly-granularity files and daily-granularity files
(auto-detected when a sheet has > 28 data rows).
"""

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from backend.utils import (
    fmt_cvr, fmt_num,
    safe_read, safe_xl,
    make_excel_styles, to_xlsx_bytes,
    xl_section, xl_header, xl_row,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

VALID_SHEETS   = {"Placed Order", "Confirmed Order", "Paid Order"}
REQUIRED_COLS  = {"Product Clicks", "Visitors", "Order Conversion Rate"}
CVR_COLS       = ["Date", "Orders", "Product Clicks", "Visitors", "Order Conversion Rate"]
SHEET_CVR_MAP  = {
    "Placed Order":    "Placed Order CVR",
    "Confirmed Order": "Confirmed Order CVR",
    "Paid Order":      "Paid Order CVR",
}
DISPLAY_HEADERS = [
    "Month / Period", "Product Clicks", "Visitors",
    "Placed Order CVR", "Confirmed Order CVR", "Paid Order CVR",
]


# ─────────────────────────────────────────────────────────────────────────────
# DATE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def parse_date(val):
    """Parse a monthly date string '01/01/2026' → datetime."""
    try:
        return val if isinstance(val, datetime) else datetime.strptime(str(val).strip(), "%d/%m/%Y")
    except Exception:
        return None


def parse_summary_date(val) -> datetime:
    """Parse the start date from a summary cell '01/01/2026-30/04/2026' → datetime."""
    try:
        return datetime.strptime(str(val).split("-")[0].strip(), "%d/%m/%Y")
    except Exception:
        return datetime.min


def to_month_label(val) -> str:
    """Convert a date value to a readable label e.g. 'January 2026'."""
    dt = parse_date(val)
    return dt.strftime("%B %Y") if dt else str(val)


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────

def validate(uploaded_file) -> tuple[bool, str]:
    """
    Returns (True, matched_sheets) when the file is a valid Shopee export.
    Returns (False, reason_string) when it is not.
    """
    if not uploaded_file.name.lower().endswith(".xlsx"):
        return False, "Expected a .xlsx file"

    try:
        xl    = safe_xl(uploaded_file)
        found = set(xl.sheet_names)
    except Exception as exc:
        return False, f"Cannot open file: {exc}"

    matched = found & VALID_SHEETS
    if not matched:
        return False, f"No Shopee sheets found — got: {found}"

    try:
        df      = safe_read(uploaded_file, sheet_name=list(matched)[0], header=None)
        headers = set(df.iloc[0].dropna().tolist())
        missing = REQUIRED_COLS - headers
        if missing:
            return False, f"Missing CVR columns: {missing}"
    except Exception as exc:
        return False, f"Error reading sheet: {exc}"

    return True, list(matched)


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTOR
# ─────────────────────────────────────────────────────────────────────────────

def extract(uploaded_file) -> dict:
    """
    Read one Shopee file and return:
      summary   — pd.Series  with combined CVR across all 3 order types
      monthly   — pd.DataFrame with one row per month
      start_date — datetime for chronological sorting
      name       — original filename
    """
    xl  = safe_xl(uploaded_file)
    raw = {}

    for sheet in SHEET_CVR_MAP:
        if sheet not in xl.sheet_names:
            continue
        df         = safe_read(uploaded_file, sheet_name=sheet, header=None)
        df.columns = df.iloc[0]
        summary    = df.iloc[1][CVR_COLS]
        monthly    = df.iloc[3:][CVR_COLS].dropna(how="all").reset_index(drop=True)
        monthly    = monthly[monthly["Date"] != "Date"].reset_index(drop=True)
        raw[sheet] = {"summary": summary, "monthly": monthly}

    base = next(
        (k for k in ("Placed Order", "Confirmed Order", "Paid Order") if k in raw),
        list(raw.keys())[0],
    )
    base_sum = raw[base]["summary"]
    base_mon = raw[base]["monthly"]

    # Daily detection: monthly rows always start on the 1st of the month.
    # If ANY row has day ≠ 1 → daily file. Handles partial months (e.g. May 1–25).
    def _is_daily(rows: pd.DataFrame) -> bool:
        for val in rows["Date"]:
            try:
                if datetime.strptime(str(val).strip(), "%d/%m/%Y").day != 1:
                    return True
            except Exception:
                pass
        return False

    # ── Parse the file's overall period start and end ──────────────────────
    def _parse_period_start(summary_date_str: str):
        try:
            return datetime.strptime(str(summary_date_str).split("-")[0].strip(), "%d/%m/%Y")
        except Exception:
            return datetime.min

    def _parse_period_end_from_summary(summary_date_str: str):
        try:
            return datetime.strptime(str(summary_date_str).split("-")[1].strip(), "%d/%m/%Y")
        except Exception:
            return datetime.min

    file_start = _parse_period_start(base_sum["Date"])
    file_end   = _parse_period_end_from_summary(base_sum["Date"])
    is_daily   = _is_daily(base_mon)   # True if rows are daily dates (day ≠ 1)

    # ── Helper: how many days does each row cover? ──────────────────────────
    def _period_days(summary_date_str: str) -> int:
        """Count days in summary period '01/05/2026-25/05/2026' → 25."""
        try:
            parts = str(summary_date_str).split("-")
            s = datetime.strptime(parts[0].strip(), "%d/%m/%Y")
            e = datetime.strptime(parts[1].strip(), "%d/%m/%Y")
            return (e - s).days + 1
        except Exception:
            return 0

    def _calendar_days(dt) -> int:
        """Days in a full calendar month."""
        import calendar
        parsed = parse_date(dt)
        return calendar.monthrange(parsed.year, parsed.month)[1] if parsed else 0

    def _month_last_day(dt) -> datetime:
        """Last day of a calendar month as datetime."""
        import calendar
        parsed = parse_date(dt)
        if not parsed:
            return datetime.min
        last = calendar.monthrange(parsed.year, parsed.month)[1]
        return datetime(parsed.year, parsed.month, last)

    if is_daily:
        # Collapse daily file to one monthly total using the summary row
        day_cnt  = _period_days(base_sum["Date"])   # e.g. 25 for May 1–25
        monthly = pd.DataFrame([{
            "Date":                file_start,
            "Paid Orders":         fmt_num(raw["Paid Order"]["summary"].get("Orders", 0)) if "Paid Order" in raw else "0",
            "Product Clicks":      fmt_num(base_sum["Product Clicks"]),
            "Visitors":            fmt_num(base_sum["Visitors"]),
            "Placed Order CVR":    fmt_cvr(raw["Placed Order"]["summary"]["Order Conversion Rate"])    if "Placed Order"    in raw else "N/A",
            "Confirmed Order CVR": fmt_cvr(raw["Confirmed Order"]["summary"]["Order Conversion Rate"]) if "Confirmed Order" in raw else "N/A",
            "Paid Order CVR":      fmt_cvr(raw["Paid Order"]["summary"]["Order Conversion Rate"])      if "Paid Order"      in raw else "N/A",
            "_day_count":          day_cnt,
            "_period_start":       file_start,
            "_period_end":         file_end,
        }])
    else:
        monthly = base_mon[["Date", "Product Clicks", "Visitors"]].copy()
        # Paid Orders count from the Paid Order sheet
        if "Paid Order" in raw:
            monthly["Paid Orders"] = raw["Paid Order"]["monthly"]["Orders"].apply(fmt_num).values
        else:
            monthly["Paid Orders"] = "0"
        for sheet, col in SHEET_CVR_MAP.items():
            if sheet in raw:
                monthly[col] = raw[sheet]["monthly"]["Order Conversion Rate"].apply(fmt_cvr).values
        monthly["Product Clicks"] = monthly["Product Clicks"].apply(fmt_num)
        monthly["Visitors"]       = monthly["Visitors"].apply(fmt_num)

        # For monthly rows: period = the calendar month
        # EXCEPT the last row, which may end before month-end if file ends mid-month
        n = len(monthly)
        p_starts = []; p_ends = []; d_counts = []
        for i, (_, row) in enumerate(monthly.iterrows()):
            p_start = parse_date(row["Date"])
            p_end   = _month_last_day(row["Date"])
            # If this is the last row AND the file ends before the month ends
            if i == n - 1 and file_end < p_end:
                p_end = file_end
            d_cnt = (p_end - p_start).days + 1 if (p_start and p_end and p_end >= p_start) else 0
            p_starts.append(p_start if p_start else datetime.min)
            p_ends.append(p_end if p_end else datetime.min)
            d_counts.append(d_cnt)

        monthly["_day_count"]    = d_counts
        monthly["_period_start"] = p_starts
        monthly["_period_end"]   = p_ends

    summary_series = pd.Series({
        "Date":                str(base_sum["Date"]),
        "Product Clicks":      fmt_num(base_sum["Product Clicks"]),
        "Visitors":            fmt_num(base_sum["Visitors"]),
        "Placed Order CVR":    fmt_cvr(raw["Placed Order"]["summary"]["Order Conversion Rate"])    if "Placed Order"    in raw else "N/A",
        "Confirmed Order CVR": fmt_cvr(raw["Confirmed Order"]["summary"]["Order Conversion Rate"]) if "Confirmed Order" in raw else "N/A",
        "Paid Order CVR":      fmt_cvr(raw["Paid Order"]["summary"]["Order Conversion Rate"])      if "Paid Order"      in raw else "N/A",
    })

    return {
        "summary":    summary_series,
        "monthly":    monthly,
        "start_date": parse_summary_date(base_sum["Date"]),
        "name":       uploaded_file.name,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(monthly_df: pd.DataFrame, summaries: list) -> bytes:
    """
    Generate a plain Shopee CVR Excel — single 'Paid Order' tab:
        Month | Paid Orders | Product Clicks | CVR   + TOTAL row
    No colors or fills. CVR read from source; TOTAL CVR = orders ÷ clicks.
    """
    from backend.utils import write_plain_sheet

    def _num(v) -> float:
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Paid Order")
    ws.sheet_view.showGridLines = True

    rows = []
    total_orders = 0.0
    total_clicks = 0.0
    for _, row in monthly_df.iterrows():
        orders = _num(row.get("Paid Orders", 0))
        clicks = _num(row.get("Product Clicks", 0))
        total_orders += orders
        total_clicks += clicks
        dt        = parse_date(row["Date"])
        month_lbl = dt.strftime("%b-%y") if dt else str(row["Date"])
        rows.append([
            month_lbl,
            f"{int(orders):,}",
            f"{int(clicks):,}",
            row.get("Paid Order CVR", "N/A"),
        ])

    total_cvr = f"{total_orders / total_clicks * 100:.2f}%" if total_clicks else "N/A"
    total_row = ["TOTAL", f"{int(total_orders):,}", f"{int(total_clicks):,}", total_cvr]

    write_plain_sheet(
        ws,
        headers=["Month", "Paid Orders", "Product Clicks", "CVR"],
        rows=rows,
        total_row=total_row,
        col_widths=[14, 14, 16, 12],
    )
    return to_xlsx_bytes(wb)
